from openpyxl import load_workbook, Workbook
import requests
import json
import os

api_key = "xxxxxxxxxxxxxxxxxxx"

if os.path.exists("result.xlsx"):
    os.remove("result.xlsx")
wb = Workbook()
source = load_workbook(filename = 'source.xlsx')
sheet_ranges = source['Лист1']
ws1 = wb.create_sheet(title="Data")
k = 1
for i in range(1,10000):
    row = f'A{i}'
    q = sheet_ranges[row].value
    if q is None:
        break
    q = q.replace(' ','%20')
    req = requests.get(f'https://api.visicom.ua/data-api/5.0/ru/geocode.json?text={q}&key={api_key}')
    data = json.loads(req.text)
    ws1.cell(column=1, row=k, value=q.replace('%20', ' '))
    if "features" in data:
        for feature in data['features']:
            ws1.cell(column=3, row=k, value=feature['geo_centroid']['coordinates'][0])
            ws1.cell(column=4, row=k, value=feature['geo_centroid']['coordinates'][1])
            k = k + 1
    else:
        try:
            ws1.cell(column=3, row=k, value=data['geo_centroid']['coordinates'][0])
            ws1.cell(column=4, row=k, value=data['geo_centroid']['coordinates'][1])
            k = k + 1
        except:
            ws1.cell(column=3, row=k, value='Не найдено')
            k = k + 1
wb.save(filename = 'result.xlsx')